import io
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.models import StationReport


def generate_excel_report(
    stations: list[StationReport],
    report_date: str,
    generated_at: datetime,
) -> io.BytesIO:
    """
    Generate an Excel report for MTA flood risk data.

    Args:
        stations: List of station reports
        report_date: Date of the report
        generated_at: Timestamp when report was generated

    Returns:
        BytesIO buffer containing the Excel file
    """
    # Convert to DataFrame
    data = []
    for station in stations:
        data.append({
            "Date": report_date,
            "Time": generated_at.strftime("%H:%M:%S"),
            "Time Zone": generated_at.tzname() if generated_at.tzinfo else "UTC",
            "Station Line": station.line,
            "Stop Name": station.station_name,
            "Borough": station.borough,
            "CBD": station.cbd,
            "Daytime Routes": station.daytime_routes,
            "Structure": station.structure,
            "GTFS Latitude": station.latitude,
            "GTFS Longitude": station.longitude,
            "Precip Rate (in/hr)": round(station.precip_rate_in_hr or 0, 3),
            "1hr Accumulation (in)": round(station.accum_1hr_in or 0, 3),
            "6hr Accumulation (in)": round(station.accum_6hr_in or 0, 3),
            "Tide Level (ft)": round(station.tide_level_ft, 2) if station.tide_level_ft else None,
            "Central Park Daily (in)": round(station.central_park_daily_in, 3)
            if station.central_park_daily_in is not None
            else None,
            "Central Park Daily Date": station.central_park_daily_date,
            "JFK Daily (in)": round(station.jfk_daily_in, 3)
            if station.jfk_daily_in is not None
            else None,
            "JFK Daily Date": station.jfk_daily_date,
            "LaGuardia Daily (in)": round(station.lga_daily_in, 3)
            if station.lga_daily_in is not None
            else None,
            "LaGuardia Daily Date": station.lga_daily_date,
            "Forecast 6hr (in)": round(station.forecast_6hr_in, 3)
            if station.forecast_6hr_in is not None
            else None,
            "Forecast 24hr (in)": round(station.forecast_24hr_in, 3)
            if station.forecast_24hr_in is not None
            else None,
            "Predicted Risk 6hr": station.predicted_risk_6hr.value
            if station.predicted_risk_6hr
            else None,
            "Predicted Risk 24hr": station.predicted_risk_24hr.value
            if station.predicted_risk_24hr
            else None,
            "Risk Level": station.risk_level.value,
            "Risk Reason": station.risk_reason or "",
            "Source": station.source,
        })

    df = pd.DataFrame(data)

    # Create Excel workbook
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Flood Risk Report", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Flood Risk Report"]

        # Style the header row
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Risk level colors
        risk_colors = {
            "FLOOD WARNING": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            "FLOOD WATCH": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
            "CLEAR": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
        }

        # Style data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1), start=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color the risk level cell (column Z = 26th column)
            risk_cell = worksheet.cell(row=row_idx, column=26)
            risk_value = risk_cell.value
            if risk_value in risk_colors:
                risk_cell.fill = risk_colors[risk_value]
                if risk_value == "FLOOD WARNING":
                    risk_cell.font = Font(color="FFFFFF", bold=True)

        # Auto-adjust column widths
        column_widths = {
            "A": 12,  # Date
            "B": 10,  # Time
            "C": 12,  # Time Zone
            "D": 12,  # Station Line
            "E": 30,  # Stop Name
            "F": 12,  # Borough
            "G": 8,   # CBD
            "H": 20,  # Daytime Routes
            "I": 12,  # Structure
            "J": 14,  # GTFS Latitude
            "K": 14,  # GTFS Longitude
            "L": 18,  # Precip Rate
            "M": 22,  # 1hr Accumulation
            "N": 22,  # 6hr Accumulation
            "O": 15,  # Tide Level
            "P": 22,  # Central Park Daily
            "Q": 20,  # Central Park Daily Date
            "R": 18,  # JFK Daily
            "S": 18,  # JFK Daily Date
            "T": 20,  # LaGuardia Daily
            "U": 20,  # LaGuardia Daily Date
            "V": 18,  # Forecast 6hr
            "W": 18,  # Forecast 24hr
            "X": 18,  # Predicted Risk 6hr
            "Y": 18,  # Predicted Risk 24hr
            "Z": 12,  # Risk Level
            "AA": 35,  # Risk Reason
            "AB": 18,  # Source
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Freeze the header row
        worksheet.freeze_panes = "A2"

        # Add summary sheet
        summary_ws = workbook.create_sheet("Summary")
        high_count = sum(1 for s in stations if s.risk_level.value == "FLOOD WARNING")
        at_risk_count = sum(1 for s in stations if s.risk_level.value == "FLOOD WATCH")
        low_count = sum(1 for s in stations if s.risk_level.value == "CLEAR")

        summary_data = [
            ["MTA Flood Risk Report Summary"],
            [""],
            ["Report Date:", report_date],
            ["Generated At:", generated_at.strftime("%Y-%m-%d %H:%M:%S UTC")],
            [""],
            ["Total Stations:", len(stations)],
            ["HIGH Risk:", high_count],
            ["AT RISK:", at_risk_count],
            ["LOW Risk:", low_count],
            [""],
            ["Data Source:", "NOAA MRMS"],
        ]

        for row_idx, row_data in enumerate(summary_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif col_idx == 1 and row_idx > 2:
                    cell.font = Font(bold=True)

        summary_ws.column_dimensions["A"].width = 20
        summary_ws.column_dimensions["B"].width = 30

    output.seek(0)
    return output


def generate_csv_report(
    stations: list[StationReport],
    report_date: str,
    generated_at: datetime,
) -> str:
    """Generate a CSV report string."""
    data = []
    for station in stations:
        data.append({
            "Date": report_date,
            "Time": generated_at.strftime("%H:%M:%S"),
            "Time Zone": generated_at.tzname() if generated_at.tzinfo else "UTC",
            "Station Line": station.line,
            "Stop Name": station.station_name,
            "Borough": station.borough,
            "CBD": station.cbd,
            "Daytime Routes": station.daytime_routes,
            "Structure": station.structure,
            "GTFS Latitude": station.latitude,
            "GTFS Longitude": station.longitude,
            "Precip Rate (in/hr)": round(station.precip_rate_in_hr or 0, 3),
            "1hr Accumulation (in)": round(station.accum_1hr_in or 0, 3),
            "6hr Accumulation (in)": round(station.accum_6hr_in or 0, 3),
            "Tide Level (ft)": round(station.tide_level_ft, 2) if station.tide_level_ft else "",
            "Central Park Daily (in)": round(station.central_park_daily_in, 3)
            if station.central_park_daily_in is not None
            else "",
            "Central Park Daily Date": station.central_park_daily_date or "",
            "JFK Daily (in)": round(station.jfk_daily_in, 3)
            if station.jfk_daily_in is not None
            else "",
            "JFK Daily Date": station.jfk_daily_date or "",
            "LaGuardia Daily (in)": round(station.lga_daily_in, 3)
            if station.lga_daily_in is not None
            else "",
            "LaGuardia Daily Date": station.lga_daily_date or "",
            "Forecast 6hr (in)": round(station.forecast_6hr_in, 3)
            if station.forecast_6hr_in is not None
            else "",
            "Forecast 24hr (in)": round(station.forecast_24hr_in, 3)
            if station.forecast_24hr_in is not None
            else "",
            "Predicted Risk 6hr": station.predicted_risk_6hr.value
            if station.predicted_risk_6hr
            else "",
            "Predicted Risk 24hr": station.predicted_risk_24hr.value
            if station.predicted_risk_24hr
            else "",
            "Risk Level": station.risk_level.value,
            "Risk Reason": station.risk_reason or "",
            "Source": station.source,
        })

    df = pd.DataFrame(data)
    return df.to_csv(index=False)
